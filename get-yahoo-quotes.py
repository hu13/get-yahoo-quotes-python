#!/usr/bin/env python

"""
get-yahoo-quotes.py:  Script to download Yahoo historical quotes using the new cookie authenticated site.

 Usage: get-yahoo-quotes SYMBOL

 History

 06-03-2017 : Created script

"""

__author__ = "Brad Luicas"
__copyright__ = "Copyright 2017, Brad Lucas"
__license__ = "MIT"
__version__ = "1.0.0"
__maintainer__ = "Brad Lucas"
__email__ = "brad@beaconhill.com"
__status__ = "Production"


import re
import sys
import time
import datetime
import requests

import progressbar
import xlsxwriter
import csv
import time
# import matplotlib.pyplot as plt, mpld3
from datetime import datetime, timedelta
import random
from urllib import urlencode
import numpy as np
from scipy.stats import norm
from sklearn.neighbors import KernelDensity
from mpl_toolkits.mplot3d import Axes3D
from constants import API_Url, Stock_Data_Key, Kernal, Stock_Val_Keys, DATE_PATTERN
from scrape_ticker_symbol import get_sp500_tickers

from openpyxl import load_workbook

def split_crumb_store(v):
    return v.split(':')[2].strip('"')


def find_crumb_store(lines):
    # Looking for
    # ,"CrumbStore":{"crumb":"9q.A4D1c.b9
    for l in lines:
        if re.findall(r'CrumbStore', l):
            return l
    print("Did not find CrumbStore")


def get_cookie_value(r):
    return {'B': r.cookies['B']}


def get_page_data(symbol):
    url = "https://finance.yahoo.com/quote/%s/?p=%s" % (symbol, symbol)
    r = requests.get(url)
    cookie = get_cookie_value(r)

    # Code to replace possible \u002F value
    # ,"CrumbStore":{"crumb":"FWP\u002F5EFll3U"
    # FWP\u002F5EFll3U
    lines = r.content.decode('unicode-escape').strip(). replace('}', '\n')
    return cookie, lines.split('\n')


def get_cookie_crumb(symbol):
    cookie, lines = get_page_data(symbol)
    crumb = split_crumb_store(find_crumb_store(lines))
    return cookie, crumb


def get_data(symbol, start_date, end_date, cookie, crumb):
    filename = '%s.csv' % (symbol)
    url = "https://query1.finance.yahoo.com/v7/finance/download/%s?period1=%s&period2=%s&interval=1d&events=history&crumb=%s" % (symbol, start_date, end_date, crumb)
    response = requests.get(url, cookies=cookie)
    with open (filename, 'wb') as handle:
        for block in response.iter_content(1024):
            handle.write(block)

def convert_date_to_epoch(date):
    return int(time.mktime(time.strptime(date, DATE_PATTERN)))

def get_now_epoch():
    # @see https://www.linuxquestions.org/questions/programming-9/python-datetime-to-epoch-4175520007/#post5244109
    return int(time.time())


def download_quotes(symbol):
    start_date = 0
    end_date = get_now_epoch()
    cookie, crumb = get_cookie_crumb(symbol)
    get_data(symbol, start_date, end_date, cookie, crumb)

def get_price_dict(num_symbols):
    return {
        Stock_Data_Key.Open: np.zeros(num_symbols),
        Stock_Data_Key.High: np.zeros(num_symbols),
        Stock_Data_Key.Low: np.zeros(num_symbols),
        Stock_Data_Key.Close: np.zeros(num_symbols),
        Stock_Data_Key.Adj_Close: np.zeros(num_symbols),
        Stock_Data_Key.Volume: np.zeros(num_symbols)
    }

def get_process_data_for_xls(symbols, only_new_update=False):
    data = {}
    index = 0
    # get data for the last 5 years
    end = get_now_epoch()
    start = int(end - timedelta(days=365*5).total_seconds())

    # widgets = ['Processed: ', progressbar.Counter('Counter: %(value)05d'),
    #            ' lines (', progressbar.Timer(), ')']

    # progressbar for the state of the download
    bar = progressbar.ProgressBar(maxval=len(symbols), \
        widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])

    bar.start()
    for symbol in symbols:
        # print 'Pulling stock data for {} ...'.format(symbol)
        if only_new_update:
            # get data for today's update
            res = get_quotes(symbol, get_now_epoch(), get_now_epoch())
        else:
            res = get_quotes(symbol, start, end)
        if res is None:
            continue

        res = res._content.split('\n')
        # skip the heading and footing
        res = res[1:len(res)-1]

        for block in res:
            points = block.split(',')
            # curr time format is YY-MM-DD
            if points[0] not in data:
                data[points[0]] = get_price_dict(len(symbols))
                price_dict = data.get(points[0])
            else:
                price_dict = data.get(points[0])

            try:
                price_dict.get(Stock_Data_Key.Open)[index] = points[1]
                price_dict.get(Stock_Data_Key.High)[index] = points[2]
                price_dict.get(Stock_Data_Key.Low)[index] = points[3]
                price_dict.get(Stock_Data_Key.Close)[index] = points[4]
                price_dict.get(Stock_Data_Key.Adj_Close)[index] = points[5]
                price_dict.get(Stock_Data_Key.Volume)[index] = points[6]

            except ValueError as e:
                print "Error: convertion error due to data missing {} in ticker {}".format(e, symbol)

        # move to the next ticker
        bar.update(index + 1)
        index += 1
    bar.finish()
    return data

def update_xls_price_file(symbols):
    # must be an int to work wit Yahoo API
    # only get the new price updates for today
    today_data = get_process_data_for_xls(symbols, True)
    try:
        # open the price_sheet
        wb = load_workbook('price_sheet.xlsx')
        # get each price type worksheet
        worksheets = {}
        last_row = 0
        for stock_dt in Stock_Val_Keys:
            worksheets[stock_dt] = wb.get_sheet_by_name(stock_dt)

        # append today's price updates to the file
        update_price_worksheet(today_data, worksheets)
        wb.save('price_sheet.xlsx')
        wb.close()
    except IOError:
        create_historical_price_sheet(symbols)

# append new updates to appropriate price_sheets
def update_price_worksheet(today_data, worksheets):
    for time in sorted(today_data):
        price_dict = today_data.get(time)
        for stock_dt in price_dict:
            wks = worksheets.get(stock_dt)

            prices = price_dict.get(stock_dt)
            wks.append([time] + prices.tolist())

def create_historical_price_sheet(symbols):
    data = get_process_data_for_xls(symbols)
    workbook = xlsxwriter.Workbook('price_sheet.xlsx')
    worksheets = {}
    for stock_dt in Stock_Val_Keys:
        worksheets[stock_dt] = workbook.add_worksheet(stock_dt)
        worksheets.get(stock_dt).write_row('A1', ['Date'] + symbols)
    # start from row 2 since row 1 is used as title
    row = 2
    for time in sorted(data):
        price_dict = data.get(time)
        for stock_dt in price_dict:
            wks = worksheets.get(stock_dt)

            prices = price_dict.get(stock_dt)

            wks.write_row('A{}'.format(row), [time] + prices.tolist())
        row += 1
    workbook.close()

def csv_write_group_by_price_type(symbols):
    price_dict = {
        Stock_Data_Key.Open: [],
        Stock_Data_Key.High: [],
        Stock_Data_Key.Low: [],
        Stock_Data_Key.Close: [],
        Stock_Data_Key.Adj_Close: [],
        Stock_Data_Key.Volume: []
    }
    longest_times = []
    for symbol in symbols:
        res = get_quotes(symbol)
        if res is None:
            continue
        quotes = process_quote_for_csv_writing(res, symbol)

        for price_type in Stock_Val_Keys:
            prices = quotes.get(price_type)
            price_dict.get(price_type).append([symbol] + prices.tolist())

        if len(quotes.get(Stock_Data_Key.Time)) > len(longest_times):
            longest_times = quotes.get(Stock_Data_Key.Time)

    if len(longest_times) == 0:
        return

    for p in price_dict:
        filename = '%s.csv' % (p)
        time_included = False
        with open (filename, 'wb') as f:
            print '------- writing %s -------' % p
            if not time_included:
                # write the date
                wr = csv.writer(f, delimiter = ',')
                wr.writerow([''] + longest_times)
                time_included = True
            # write the price
            wr.writerows(price_dict.get(p))
            print '--------------------------'

def download_many_quotes(symbols):
    assert type(symbols) is list
    for symbol in symbols:
        print("--------------------------------------------------")
        print("Downloading %s to %s.csv" % (symbol, symbol))
        try:
            download_quotes(symbol)
            print("--------------------------------------------------")
        except KeyError:
            print("Invalid stock ticker!")


# by default,
# start_date = the beginning of the historical data available
# end_date = now
# frequency = daily
def get_quotes(symbol, start_date=0, end_date=get_now_epoch(), frequency='1d'):
    try:
        cookie, crumb = get_cookie_crumb(symbol)
        return request_data(symbol, start_date, end_date, frequency, cookie, crumb)
    except KeyError:
        print("Invalid stock ticker {}!".format(symbol))
        return None

def request_data(symbol, start_date, end_date, frequency, cookie, crumb):
    url = "%s%s?period1=%s&period2=%s&interval=%s&events=history&crumb=%s" % (API_Url.historical_quotes, symbol, start_date, end_date, frequency, crumb)
    return requests.get(url, cookies=cookie)

def process_data(response):
    # Date,Open,High,Low,Close,Adj Close,Volume
    data = {}
    response = response._content.split('\n')
    # skip the heading and footing
    response = response[1:len(response)-1]
    for block in response:
        points = block.split(',')
        time = convert_date_to_sec(points[0])
        data[points[0]] = {
            Stock_Data_Key.Time: time,
            Stock_Data_Key.Open: points[1],
            Stock_Data_Key.High: points[2],
            Stock_Data_Key.Low: points[3],
            Stock_Data_Key.Close: points[4],
            Stock_Data_Key.Adj_Close: points[5],
            Stock_Data_Key.Volume: points[6]
        }
    return data

def ignore_null_data(val):
    if type(val) is not float:
        return 0
    return val

# put all quotes of this stock into separate np array
# group by types of stock price period
# order by latest first
def process_quote_for_csv_writing(response, symbol):
    response = response._content.split('\n')
    # skip the heading and footing
    response = response[1:len(response)-1]

    num_data_points = len(response)
    stock_data = {
        Stock_Data_Key.Time: [],
        Stock_Data_Key.Open: np.empty(num_data_points),
        Stock_Data_Key.High: np.empty(num_data_points),
        Stock_Data_Key.Low: np.empty(num_data_points),
        Stock_Data_Key.Close: np.empty(num_data_points),
        Stock_Data_Key.Adj_Close: np.empty(num_data_points),
        Stock_Data_Key.Volume: np.empty(num_data_points)
    }

    index = 0
    for block in reversed(response):
        points = block.split(',')
        # points = [ignore_null_data(p) for p in points]
        # curr time format is YY-MM-DD
        # time = convert_date_to_sec(points[0])
        try:
            stock_data.get(Stock_Data_Key.Time).append(points[0])
            stock_data.get(Stock_Data_Key.Open)[index] = points[1]
            stock_data.get(Stock_Data_Key.High)[index] = points[2]
            stock_data.get(Stock_Data_Key.Low)[index] = points[3]
            stock_data.get(Stock_Data_Key.Close)[index] = points[4]
            stock_data.get(Stock_Data_Key.Adj_Close)[index] = points[5]
            stock_data.get(Stock_Data_Key.Volume)[index] = points[6]
            index += 1
        except ValueError as e:
            print "Error: convertion error due to data missing {} in ticker {}".format(e, symbol)
    return stock_data

def get_quote(code, historical_quotes):
    if code not in Stock_Data_Key.__dict__.values():
        raise AttributeError()
    if type(historical_quotes) is not dict:
        raise TypeError()

    times = np.empty(len(historical_quotes.keys()))
    quotes = np.empty(len(historical_quotes.keys()))
    index = 0

    for time in sorted(historical_quotes.keys(), reverse=True):
        times[index] = historical_quotes.get(time).get(Stock_Data_Key.Time)
        quotes[index] = float(historical_quotes.get(time).get(code))
        index += 1
    return times, quotes

def update_title(axes, curr_data_set):
    axes.set_title(datetime.now())
    # axes.plot(curr_data_set, 'ks-', mec='w')
    axes.figure.canvas.draw()


def convert_date_to_sec(date):
    parts = date.split('-')
    assert (len(parts) == 3)
    t = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
    return (t-datetime(1970, 1, 1)).total_seconds()

#FIXME: need to understand how KDE works
def KDE_for_pdf(price, kernal_arg=Kernal.Epanechnikov, bandwidth_val=2):
    points = [[p] for p in price]
    kde = KernelDensity(kernel=kernal_arg, bandwidth=bandwidth_val).fit(points)
    log_dens = kde.score_samples(points)
    return np.exp(log_dens)

def surface_plots(datasets):
    fig, ax = plt.subplots(subplot_kw={'projection': '3d'})
    for dataset in datasets:
        ax.plot(dataset.get('time'), dataset.get('price'), dataset.get('prob'))
    ax.set_xlabel('time in sec')
    ax.set_ylabel('stock price')
    ax.set_zlabel('estimated prob from KDE')

# if __name__ == '__main__':
#     # If we have at least one parameter go ahead and loop overa all the parameters assuming they are symbols
#     response = get_quotes('vz', frequency='1wk')
#     historical_quotes = process_data(response)
#     times, high = get_quote(Stock_Data_Key.High, historical_quotes)
#
#     # ------------------------------- plot of time vs stock ------------------------------- #
#
#     plt.subplot(311)
#     plt.scatter(times, high, marker='o', s=1.0)
#
#     # timer = fig.canvas.new_timer(interval=100)
#     # timer.add_callback(update_title, my_plot, i[rn.randint(0,1)])
#     # timer.start()
#     # plt.show()
#
#     # ------------------------------- end of plot of time vs stock ------------------------------- #
#
#     # histogram of High price
#     plt.subplot(312)
#     plt.hist(high)
#
#     data_sets = []
#     for end_point in [len(high), 1, len(high)/2, len(high)/3]:
#         density_vals = KDE_for_pdf(high[0:end_point])
#
#         # print density_vals
#         # print high[0:end_point]
#
#         data = {'time': [i for i in range(0, end_point)], 'price': high[0:end_point], 'prob': density_vals}
#         data_sets.append(data)
#     surface_plots(data_sets)
#     plt.show()

if __name__ == '__main__':
    # If we have at least one parameter go ahead and loop overa all the parameters assuming they are symbols
    if len(sys.argv) == 1:
        print("\nUsage: get-yahoo-quotes.py SYMBOL\n\n")
        # symbols = get_sp500_tickers()
        # csv_write_group_by_price_type(symbols)
    else:
        # download_many_quotes(sys.argv[1:])
        if len(sys.argv) == 3:
            if sys.argv[1] == '--file':
                tickers = []
                with open(sys.argv[2], 'r') as f:
                    for line in f:
                        line = line.split(',')[0]
                        tickers.append(line)
                # csv_write_group_by_price_type(tickers)
                update_xls_price_file(tickers)
            else:
                # csv_write_group_by_price_type(sys.argv[1:])
                update_xls_price_file(sys.argv[1:])
        else:
            # csv_write_group_by_price_type(sys.argv[1:])
            update_xls_price_file(sys.argv[1:])
